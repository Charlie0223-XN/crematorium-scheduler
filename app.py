# app.py
from flask import Flask, request, jsonify, render_template, send_file
from scheduler import generate_day, generate_period, EMPLOYEES
from datetime import datetime

import io
from openpyxl import Workbook

app = Flask(__name__)


# -----------------------------
# 首頁：多日排班 UI
# -----------------------------
@app.route("/")
def index():
    return render_template("index.html", employees=EMPLOYEES)


# -----------------------------
# 單日排班 API（保留測試用）
# -----------------------------
@app.route("/api/schedule", methods=["POST"])
def api_schedule():
    data = request.get_json()
    employees = data.get("employees", [])
    prev_day = data.get("prev_day", {})

    if not employees:
        return jsonify({"error": "請至少選一個上班人員"}), 400

    assignment, score = generate_day(employees, prev_day)
    return jsonify({"assignment": assignment, "score": score})


# -----------------------------
# 多日排班 API（主力）
# -----------------------------
@app.route("/api/schedule_range", methods=["POST"])
def api_schedule_range():
    """
    接收多天排班需求。

    request JSON 格式預期為：
    {
      "days": [
        {
          "date": "2025-12-01",
          "employees": ["豐杰", "在慶"],
          "full_staff": false,
          "big_day": true
        },
        ...
      ]
    }
    """
    data = request.get_json()
    days = data.get("days", [])

    if not days:
        return jsonify({"error": "至少要提供一天的資料"}), 400

    days_info = []
    for day in days:
        full_staff = day.get("full_staff", False)
        big_day = day.get("big_day", False)
        date_str = day.get("date")

        if not date_str:
            return jsonify({"error": "缺少日期資訊"}), 400

        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": f"日期格式錯誤：{date_str}"}), 400

        weekday = dt.weekday()  # Monday=0 ... Sunday=6

        if full_staff:
            emps = EMPLOYEES[:]  # 全員到齊
        else:
            emps = day.get("employees", [])
            if not emps:
                return jsonify({"error": f"{date_str} 沒有勾任何上班人，且未勾 full_staff"}), 400

        days_info.append({
            "date": date_str,
            "weekday": weekday,
            "big_day": bool(big_day),
            "employees": emps,
        })

    # 用你最後版的演算法排整段
    schedule = generate_period(days_info)

    # 回傳時把 date / big_day 一起帶回去，方便前端顯示與統計
    result = []
    for idx, (meta, assign) in enumerate(zip(days_info, schedule), start=1):
        result.append({
            "day_index": idx,
            "date": meta["date"],
            "big_day": bool(meta.get("big_day", False)),
            "assignment": assign
        })

    return jsonify({"schedule": result})



# -----------------------------
# 匯出 Excel API
# -----------------------------
@app.route("/api/export_excel", methods=["POST"])
def export_excel():
    """
    期待前端傳來：
    {
      "schedule": [
        { "day_index": 1, "date": "...", "assignment": {...}, ... },
        ...
      ],
      "start_date": "2024-11-09",
      "end_date": "2024-12-06"
    }
    """
    data = request.get_json()
    schedule = data.get("schedule", [])
    start_date = data.get("start_date", "")
    end_date = data.get("end_date", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    row = 1
    for day in schedule:
        day_index = day.get("day_index")
        date_str = day.get("date", "")
        assignment = day.get("assignment", {}) or {}

        # Day n + 日期
        if date_str:
            ws.cell(row=row, column=1, value=f"Day {day_index}  ({date_str})")
        else:
            ws.cell(row=row, column=1, value=f"Day {day_index}")
        row += 1

        # 依姓名排序
        for name in sorted(assignment.keys()):
            role = assignment[name]
            ws.cell(row=row, column=1, value=f"{name}：{role}")
            row += 1

        # 空一行
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if start_date and end_date:
        filename = f"schedule_{start_date}_to_{end_date}.xlsx"
    else:
        filename = "schedule.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
